"""
Copyright 2023 binary butterfly GmbH
Use of this source code is governed by an MIT-style license that can be found in the LICENSE.txt.
"""

from io import BytesIO, StringIO
from zipfile import BadZipFile

from lxml import etree
from lxml.etree import ParseError
from openpyxl.reader.excel import load_workbook
from parkapi_sources.converters.base_converter.push import CsvConverter, JsonConverter, XlsxConverter, XmlConverter
from parkapi_sources.exceptions import ImportParkingSiteException
from parkapi_sources.models import RealtimeParkingSiteInput, StaticParkingSiteInput

from webapp.admin_rest_api import AdminApiBaseHandler
from webapp.common.rest.exceptions import InvalidInputException
from webapp.models import Source
from webapp.repositories import ParkingSiteRepository, SourceRepository
from webapp.services.import_service import ParkingSiteGenericImportService


class GenericParkingSitesHandler(AdminApiBaseHandler):
    source_repository: SourceRepository
    parking_site_repository: ParkingSiteRepository
    parking_site_generic_import_service: ParkingSiteGenericImportService

    def __init__(
        self,
        *args,
        source_repository: SourceRepository,
        parking_site_repository: ParkingSiteRepository,
        parking_site_generic_import_service: ParkingSiteGenericImportService,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)
        self.source_repository = source_repository
        self.parking_site_repository = parking_site_repository
        self.parking_site_generic_import_service = parking_site_generic_import_service

    def handle_json_data(
        self,
        source_uid: str,
        data: dict | list,
    ) -> tuple[list[StaticParkingSiteInput | RealtimeParkingSiteInput], list[ImportParkingSiteException]]:
        source = self.parking_site_generic_import_service.get_upserted_source(source_uid)
        import_service: JsonConverter = self.parking_site_generic_import_service.park_api_sources.converter_by_uid[source_uid]  # type: ignore

        parking_site_inputs, parking_site_errors = import_service.handle_json(data)

        self._handle_import_results(source, parking_site_inputs, parking_site_errors)

        return parking_site_inputs, parking_site_errors

    def handle_xml_data(
        self, source_uid: str, data: bytes
    ) -> tuple[list[StaticParkingSiteInput | RealtimeParkingSiteInput], list[ImportParkingSiteException]]:
        source = self.parking_site_generic_import_service.get_upserted_source(source_uid)
        import_service: XmlConverter = self.parking_site_generic_import_service.park_api_sources.converter_by_uid[source_uid]  # type: ignore

        try:
            root_element = etree.fromstring(data, parser=etree.XMLParser(resolve_entities=False))  # noqa: S320
        except ParseError as e:
            raise InvalidInputException(message='Invalid XML file') from e

        parking_site_inputs, parking_site_errors = import_service.handle_xml(root_element)

        self._handle_import_results(source, parking_site_inputs, parking_site_errors)

        return parking_site_inputs, parking_site_errors

    def handle_csv_data(
        self, source_uid: str, data: str
    ) -> tuple[list[StaticParkingSiteInput | RealtimeParkingSiteInput], list[ImportParkingSiteException]]:
        source = self.parking_site_generic_import_service.get_upserted_source(source_uid)
        import_service: CsvConverter = self.parking_site_generic_import_service.park_api_sources.converter_by_uid[source_uid]  # type: ignore

        try:
            parking_site_inputs, parking_site_errors = import_service.handle_csv_string(StringIO(data))
        except Exception as e:
            raise InvalidInputException(message=f'Invalid input: {getattr(e, "message", "unknown reason")}') from e

        self._handle_import_results(source, parking_site_inputs, parking_site_errors)

        return parking_site_inputs, parking_site_errors

    def handle_xlsx_data(
        self,
        source_uid: str,
        data: bytes,
    ) -> tuple[list[StaticParkingSiteInput | RealtimeParkingSiteInput], list[ImportParkingSiteException]]:
        source = self.parking_site_generic_import_service.get_upserted_source(source_uid)
        import_service: XlsxConverter = self.parking_site_generic_import_service.park_api_sources.converter_by_uid[source_uid]  # type: ignore

        try:
            workbook = load_workbook(filename=BytesIO(data))
        # Sadly, there is no generic parent exception load_workbook throws, so this list might be incomplete
        except (BadZipFile, KeyError, ValueError) as e:
            raise InvalidInputException(message='Invalid XLSX file') from e

        try:
            parking_site_inputs, parking_site_errors = import_service.handle_xlsx(workbook)
        except Exception as e:
            raise InvalidInputException(message=f'Invalid input: {getattr(e, "message", "unknown reason")}') from e

        self._handle_import_results(source, parking_site_inputs, parking_site_errors)

        return parking_site_inputs, parking_site_errors

    def _handle_import_results(
        self,
        source: Source,
        parking_site_inputs: list[StaticParkingSiteInput | RealtimeParkingSiteInput],
        parking_site_errors: list[ImportParkingSiteException],
    ):
        # TODO: it might make sense to support combined inputs, depending on which data types we get in future. So far, we split it up, and
        #  have the issue that the errors cannot be assigned to static or realtime
        static_parking_site_inputs = [item for item in parking_site_inputs if isinstance(item, StaticParkingSiteInput)]

        if len(static_parking_site_inputs):
            self.parking_site_generic_import_service.handle_static_import_results(source, static_parking_site_inputs, parking_site_errors)

        realtime_parking_site_inputs = [item for item in parking_site_inputs if isinstance(item, RealtimeParkingSiteInput)]

        if len(realtime_parking_site_inputs):
            self.parking_site_generic_import_service.handle_realtime_import_results(
                source, realtime_parking_site_inputs, parking_site_errors
            )

        self.parking_site_repository.commit_transaction()

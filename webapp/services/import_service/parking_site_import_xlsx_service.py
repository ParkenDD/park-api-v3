"""
Copyright 2023 binary butterfly GmbH
Use of this source code is governed by an MIT-style license that can be found in the LICENSE.txt.
"""

from datetime import datetime, timezone
from pathlib import Path
from random import randint
from typing import List

from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from validataclass.exceptions import ValidationError
from validataclass.validators import DataclassValidator

from webapp.models import ParkingSite, Source
from webapp.models.parking_site import ParkingSiteType
from webapp.repositories import ParkingSiteRepository, SourceRepository
from webapp.repositories.exceptions import ObjectNotFoundException
from webapp.services.base_service import BaseService
from webapp.services.import_service.exceptions import (
    ImportDatasetException,
    ImportException,
)
from webapp.services.import_service.parking_site_import_xlsx_validator import (
    ParkingSiteInput,
)


class ParkingSiteXlsxImportService(BaseService):
    parking_site_repository: ParkingSiteRepository
    source_repository: SourceRepository

    parking_site_validator = DataclassValidator(ParkingSiteInput)

    header_row: dict[str, str] = {
        'ID': 'original_uid',
        'Name': 'name',
        'Art der Anlage': 'type',
        'Betreiber Name': 'operator_name',
        'Längengrad': 'lat',
        'Breitengrad': 'lon',
        'Adresse mit PLZ und Stadt': 'address',
        'Maximale Parkdauer': 'max_stay',
        'Anzahl Stellplätze': 'capacity',
        'Anzahl Carsharing-Parkplätze': 'capacity_carsharing',
        'Anzahl Ladeplätze': 'capacity_charging',
        'Anzahl Frauenparkplätze': 'capacity_woman',
        'Anzahl Behindertenparkplätze': 'capacity_disabled',
        'Anlage beleuchtet?': 'has_lighting',
        'gebührenpflichtig?': 'has_fee',
        'Existieren Live-Daten?': 'has_realtime_data',
        'Gebühren-Informationen': 'fee_description',
        '24/7 geöffnet?': 'opening_hours_is_24_7',
        'Öffnungszeiten Mo-Fr Beginn': 'opening_hours_weekday_begin',
        'Öffnungszeiten Mo-Fr Ende': 'opening_hours_weekday_end',
        'Öffnungszeiten Sa Beginn': 'opening_hours_saturday_begin',
        'Öffnungszeiten Sa Ende': 'opening_hours_saturday_end',
        'Öffnungszeiten So Beginn': 'opening_hours_sunday_begin',
        'Öffnungszeiten So Ende': 'opening_hours_sunday_end',
        'Webseite': 'public_url',
        'Park&Ride': 'is_park_and_ride',
        'Weitere öffentliche Informationen': 'description',
    }
    type_mapping: dict[str, ParkingSiteType] = {
        'Parkplatz': ParkingSiteType.OFF_STREET_PARKING_GROUND,
        'Parkhaus': ParkingSiteType.CAR_PARK,
        'Tiefgarage': ParkingSiteType.UNDERGROUND,
        'Am Straßenrand': ParkingSiteType.ON_STREET,
    }

    def __init__(
        self,
        *args,
        parking_site_repository: ParkingSiteRepository,
        source_repository: SourceRepository,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)
        self.parking_site_repository = parking_site_repository
        self.source_repository = source_repository

    def load_and_import_parking_sites(self, source_uid: str, import_file_path: Path):
        try:
            source = self.source_repository.fetch_source_by_uid(source_uid)
        except ObjectNotFoundException:
            source = Source()
            source.uid = source_uid
            self.source_repository.save_source(source)

        worksheet = self.load_parking_sites(import_file_path)

        mapping = self.get_mapping_by_header(next(worksheet.rows))

        self.import_parking_sites(
            source=source,
            worksheet=worksheet,
            mapping=mapping,
        )

        source.last_import = datetime.now(tz=timezone.utc)
        self.source_repository.save_source(source)

    @staticmethod
    def load_parking_sites(import_file_path: Path) -> Worksheet:
        return load_workbook(filename=str(import_file_path)).active

    def get_mapping_by_header(self, row: tuple[Cell]) -> list[str]:
        header_keys = self.header_row.keys()
        mapping: list[str] = []
        for col in row:
            if col.value not in header_keys:
                raise ImportException(message=f'invalid header column {col.value}')
            mapping.append(self.header_row[col.value])
        return mapping

    def import_parking_sites(self, source: Source, worksheet: Worksheet, mapping: list[str]) -> List[ImportDatasetException]:
        validation_exceptions: List[ImportDatasetException] = []
        for row in worksheet.iter_rows(min_row=2):
            parking_site_dict = {}
            for position, field in enumerate(mapping):
                parking_site_dict[field] = row[position].value
            try:
                parking_site_input = self.parking_site_validator.validate(parking_site_dict)
            except ValidationError as e:
                validation_exceptions.append(ImportDatasetException(exception=e, dataset=parking_site_dict))
                continue

            try:
                parking_site = self.parking_site_repository.fetch_parking_site_by_source_id_and_external_uid(
                    source_id=source.id,
                    original_uid=parking_site_input.original_uid,
                )
            except ObjectNotFoundException:
                parking_site = ParkingSite()
                parking_site.source_id = source.id
                parking_site.original_uid = parking_site_input.original_uid

            direct_copy_keys = [
                'name',
                'operator_name',
                'lat',
                'lon',
                'address',
                'max_stay',
                'capacity',
                'capacity_carsharing',
                'capacity_charging',
                'capacity_charging',
                'capacity_woman',
                'capacity_disabled',
                'has_lighting',
                'has_fee',
                'has_live_data',
                'fee_description',
                'public_url',
                'is_park_ride',
                'description',
                'is_supervised',
            ]
            for key in direct_copy_keys:
                setattr(parking_site, key, getattr(parking_site_input, key))

            parking_site.type = self.type_mapping.get(parking_site_input.type)

            if parking_site_input.opening_hours_is_24_7 is True:
                parking_site.opening_hours = '24/7'
            else:
                # TODO: opening hours over midnight
                opening_hours_fragments = []
                if parking_site_input.opening_hours_weekday_begin and parking_site_input.opening_hours_weekday_end:
                    opening_hours_fragments.append(
                        f'Mo-Fr {parking_site_input.opening_hours_weekday_begin.strftime("%H:%M")}'
                        f'-{parking_site_input.opening_hours_weekday_end.strftime("%H:%M")}',
                    )
                if parking_site_input.opening_hours_saturday_begin and parking_site_input.opening_hours_saturday_end:
                    opening_hours_fragments.append(
                        f'Sa {parking_site_input.opening_hours_saturday_begin.strftime("%H:%M")}'
                        f'-{parking_site_input.opening_hours_saturday_end.strftime("%H:%M")}',
                    )
                if parking_site_input.opening_hours_sunday_begin and parking_site_input.opening_hours_sunday_end:
                    opening_hours_fragments.append(
                        f'Su {parking_site_input.opening_hours_sunday_begin.strftime("%H:%M")}'
                        f'-{parking_site_input.opening_hours_sunday_end.strftime("%H:%M")}',
                    )
                parking_site.opening_hours = '; '.join(opening_hours_fragments)

            parking_site.static_data_updated_at = datetime.now(tz=timezone.utc)
            self.parking_site_repository.save_parking_site(parking_site)

        return validation_exceptions
